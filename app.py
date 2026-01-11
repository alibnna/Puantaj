"""
Profesyonel Puantaj ve Alacak Hesaplama Sistemi
================================================
Bordro okuma ve finansal hesaplama - ORİJİNAL NOTEBOOK MANTIĞI
"""

import pandas as pd
import numpy as np
import os
import glob
import re
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import streamlit as st
from streamlit.runtime.scriptrunner import get_script_run_ctx
import os
import shutil

import time

# ============================================================================
# 1. KONFIGURASYON SINIFI
# ============================================================================
class Config:
    """Yolları çalışma dizinine göre dinamik hale getiriyoruz"""
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    BORDRO_KLASORU = os.path.join(BASE_DIR, "Bordro")
    BORDRO_PDF = os.path.join(BORDRO_KLASORU, "islenen_bordro.pdf")
    
    HAM_VERI_KLASORU = os.path.join(BASE_DIR, "HamVeriler")
    TEMIZ_VERI_KLASORU = os.path.join(BASE_DIR, "PDKS")
    PUANTAJ_KLASORU = os.path.join(BASE_DIR, "Olusturulan_Puantajlar")
    RAPOR_KLASORU = os.path.join(BASE_DIR, "Final_Hesaplama")
    
    TEMIZ_VERI_DOSYASI = "Birlestirilmis_Temiz_Veri.xlsx"
    ALACAK_RAPORU_DOSYASI = "HASSAS_ALACAK_RAPORU.xlsx"

    # İş Kanunu Sabitleri
    GUNLUK_STANDART_SAAT = 7.5
    HAFTALIK_YASAL_SURE = 45.0
    MINIMUM_SURE_GARANTISI = True
    HAFTA_TATILI_ISARETI = "x"

    GECE_CALISMA_TOLERANS_DK = 0 # Varsayılan 0 (Kapalı)

    # Excel Renk Kodları
    RENK_SARI = "FFC000"
    RENK_GRI = "D9D9D9"
    RENK_YESIL = "92D050"
    RENK_FOSFOR = "FFFF00"
    RENK_KIRMIZI = "000000"
    RENK_MAVI = "002060"
    RENK_KOYU_KIRMIZI = "C00000"
    RENK_KOYU_MAVI = "0070C0"

    # Gün İsimleri
    GUN_ISIMLERI = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']

    # Ay İsimleri (Bordro için)
    AYLAR = {
        'Ocak': 1, 'Şubat': 2, 'Mart': 3, 'Nisan': 4, 'Mayıs': 5, 'Haziran': 6,
        'Temmuz': 7, 'Ağustos': 8, 'Eylül': 9, 'Ekim': 10, 'Kasım': 11, 'Aralık': 12
    }

    @staticmethod
    def get_user_base():
        # Her kullanıcıya özel benzersiz Session ID alıyoruz
        ctx = get_script_run_ctx()
        session_id = ctx.session_id if ctx else "default_user"
        user_path = os.path.join(Config.BASE_DIR, "UserData", session_id)
        return user_path

    # Yollar artık fonksiyonel olarak kullanıcı bazlı oluşturuluyor
    @staticmethod
    def get_paths():
        # Benzersiz Kullanıcı Kimliği (Session ID)
        ctx = get_script_run_ctx()
        session_id = ctx.session_id if ctx else "default_user"
        
        # Kullanıcıya özel ana klasör
        user_base = os.path.join(Config.BASE_DIR, "UserData", session_id)
        
        # Tüm alt klasör yolları
        paths = {
            "BASE": user_base,
            "HAM": os.path.join(user_base, "HamVeriler"),
            "BORDRO": os.path.join(user_base, "Bordro"),
            "PDKS": os.path.join(user_base, "PDKS"),
            "PUANTAJ": os.path.join(user_base, "Olusturulan_Puantajlar"),
            "RAPOR": os.path.join(user_base, "Final_Hesaplama")
        }
        # İşlenen PDF dosyasının tam yolu
        paths["BORDRO_FILE"] = os.path.join(paths["BORDRO"], "islenen_bordro.pdf")
        return paths

    @staticmethod
    def klasorleri_hazirla():
        paths = Config.get_paths()
        # Sözlükteki tüm klasör yollarını tek tek kontrol et ve oluştur
        for key, path in paths.items():
            if key != "BORDRO_FILE": # Dosya yolunu klasör gibi oluşturmaya çalışma
                os.makedirs(path, exist_ok=True)
        return paths

# ============================================================================
# 2. STATİK TATİL VERİTABANI (2000-2025)
# ============================================================================
class StaticHolidays:
    """
    2000-2025 yılları arası tüm resmi tatil, bayram ve arefeleri hardcode içerir.
    """

    @staticmethod
    def _tatil_veritabani_olustur():
        """Tüm tatil ve arefeleri içeren dict oluşturur"""
        tatiller = {}
        arefeler = set()

        sabit_tatiller = {
            (1, 1): "Yılbaşı",
            (4, 23): "23 Nisan",
            (5, 1): "1 Mayıs",
            (5, 19): "19 Mayıs",
            (7, 15): "15 Temmuz",
            (8, 30): "30 Ağustos",
            (10, 29): "29 Ekim"
        }

        ramazan_bayrami = {
            2000: [(12, 27), (12, 28), (12, 29)], 2001: [(12, 16), (12, 17), (12, 18)],
            2002: [(12, 5), (12, 6), (12, 7)], 2003: [(11, 25), (11, 26), (11, 27)],
            2004: [(11, 14), (11, 15), (11, 16)], 2005: [(11, 3), (11, 4), (11, 5)],
            2006: [(10, 23), (10, 24), (10, 25)], 2007: [(10, 13), (10, 14), (10, 15)],
            2008: [(9, 30), (10, 1), (10, 2)], 2009: [(9, 20), (9, 21), (9, 22)],
            2010: [(9, 10), (9, 11), (9, 12)], 2011: [(8, 30), (8, 31), (9, 1)],
            2012: [(8, 19), (8, 20), (8, 21)], 2013: [(8, 8), (8, 9), (8, 10)],
            2014: [(7, 28), (7, 29), (7, 30)], 2015: [(7, 17), (7, 18), (7, 19)],
            2016: [(7, 5), (7, 6), (7, 7)], 2017: [(6, 25), (6, 26), (6, 27)],
            2018: [(6, 15), (6, 16), (6, 17)], 2019: [(6, 4), (6, 5), (6, 6)],
            2020: [(5, 24), (5, 25), (5, 26)], 2021: [(5, 13), (5, 14), (5, 15)],
            2022: [(5, 2), (5, 3), (5, 4)], 2023: [(4, 21), (4, 22), (4, 23)],
            2024: [(4, 10), (4, 11), (4, 12)], 2025: [(3, 30), (3, 31), (4, 1)]
        }

        kurban_bayrami = {
            2000: [(3, 16), (3, 17), (3, 18), (3, 19)], 2001: [(3, 5), (3, 6), (3, 7), (3, 8)],
            2002: [(2, 23), (2, 24), (2, 25), (2, 26)], 2003: [(2, 11), (2, 12), (2, 13), (2, 14)],
            2004: [(2, 1), (2, 2), (2, 3), (2, 4)], 2005: [(1, 21), (1, 22), (1, 23), (1, 24)],
            2006: [(1, 10), (1, 11), (1, 12), (1, 13)], 2007: [(12, 20), (12, 21), (12, 22), (12, 23)],
            2008: [(12, 8), (12, 9), (12, 10), (12, 11)], 2009: [(11, 27), (11, 28), (11, 29), (11, 30)],
            2010: [(11, 16), (11, 17), (11, 18), (11, 19)], 2011: [(11, 6), (11, 7), (11, 8), (11, 9)],
            2012: [(10, 25), (10, 26), (10, 27), (10, 28)], 2013: [(10, 15), (10, 16), (10, 17), (10, 18)],
            2014: [(10, 4), (10, 5), (10, 6), (10, 7)], 2015: [(9, 24), (9, 25), (9, 26), (9, 27)],
            2016: [(9, 12), (9, 13), (9, 14), (9, 15)], 2017: [(9, 1), (9, 2), (9, 3), (9, 4)],
            2018: [(8, 21), (8, 22), (8, 23), (8, 24)], 2019: [(8, 11), (8, 12), (8, 13), (8, 14)],
            2020: [(7, 31), (8, 1), (8, 2), (8, 3)], 2021: [(7, 20), (7, 21), (7, 22), (7, 23)],
            2022: [(7, 9), (7, 10), (7, 11), (7, 12)], 2023: [(6, 28), (6, 29), (6, 30), (7, 1)],
            2024: [(6, 16), (6, 17), (6, 18), (6, 19)], 2025: [(6, 6), (6, 7), (6, 8), (6, 9)]
        }

        for yil in range(2000, 2026):
            for (ay, gun), isim in sabit_tatiller.items():
                tarih = datetime(yil, ay, gun).date()
                tatiller[tarih] = isim

            arefe_28_ekim = datetime(yil, 10, 28).date()
            arefeler.add(arefe_28_ekim)

            if yil in ramazan_bayrami:
                for ay, gun in ramazan_bayrami[yil]:
                    tarih = datetime(yil, ay, gun).date()
                    tatiller[tarih] = "Ramazan Bayramı"
                ilk_gun_ay, ilk_gun_gun = ramazan_bayrami[yil][0]
                arefe = datetime(yil, ilk_gun_ay, ilk_gun_gun).date() - timedelta(days=1)
                arefeler.add(arefe)

            if yil in kurban_bayrami:
                for ay, gun in kurban_bayrami[yil]:
                    tarih = datetime(yil, ay, gun).date()
                    tatiller[tarih] = "Kurban Bayramı"
                ilk_gun_ay, ilk_gun_gun = kurban_bayrami[yil][0]
                arefe = datetime(yil, ilk_gun_ay, ilk_gun_gun).date() - timedelta(days=1)
                arefeler.add(arefe)

        return tatiller, arefeler

    _TATIL_DB, _AREFE_SET = _tatil_veritabani_olustur()

    @classmethod
    def get_status(cls, tarih):
        if not isinstance(tarih, datetime):
            tarih = tarih.date() if hasattr(tarih, 'date') else tarih
        if tarih in cls._AREFE_SET:
            return 'AREFE'
        if tarih in cls._TATIL_DB:
            return 'UBGT'
        return 'NORMAL'

    @classmethod
    def is_arefe(cls, tarih):
        if not isinstance(tarih, datetime):
            tarih = tarih.date() if hasattr(tarih, 'date') else tarih
        return tarih in cls._AREFE_SET

    @classmethod
    def is_ubgt(cls, tarih):
        if not isinstance(tarih, datetime):
            tarih = tarih.date() if hasattr(tarih, 'date') else tarih
        return tarih in cls._TATIL_DB


# ============================================================================
# 3. ETL İŞÇİSİ
# ============================================================================
class ETLWorker:
    """Ham verileri okur, bozuk karakterleri (İsim, Gün, PG) ve saatleri temizler"""

    # Gelişmiş Karakter Haritası
    KARAKTER_HARITASI = {
        'Ã„°': 'İ', 'Ã': 'Ğ', 'Ãž': 'Ş', 'Ã¾': 'ş', 'Ã½': 'ı', 'Ã°': 'ğ',
        'Ý': 'İ', 'Ð': 'Ğ', 'Þ': 'Ş', 'þ': 'ş', 'ý': 'ı', 'ð': 'ğ',
        'Ä°': 'İ', 'Äž': 'Ğ', 'Åž': 'Ş', 'ÅŸ': 'ş', 'Ä±': 'ı', 'ÄŸ': 'ğ',
        'Ã‡': 'Ç', 'Ã§': 'ç', 'Ã–': 'Ö', 'Ã¶': 'ö', 'Ãœ': 'Ü', 'Ã¼': 'ü',
        'Ý': 'İ', 'Þ': 'Ş', 'Ð': 'Ğ', # Ekstra garantiler
        'GÜNDÜZ': 'GÜNDÜZ' # Bazen GÜNDÜZ kelimesi bozulmaz ama referans olsun
    }

    @classmethod
    def turkce_karakter_duzelt(cls, text):
        """Metindeki bozuk karakterleri düzeltir"""
        if not isinstance(text, str):
            return text
        
        # Önce haritadaki bilinen bozuklukları düzelt
        for bozuk, duzgun in cls.KARAKTER_HARITASI.items():
            text = text.replace(bozuk, duzgun)
            
        # Ekstra temizlik: Bazen tek kalan bozuk harfler olabilir
        return text.strip()

    @classmethod
    def saat_formatla(cls, val):
        """Veriyi 08:00 formatında temiz stringe çevirir"""
        val_str = str(val).strip()
        if val_str.lower() in ['nan', '', 'nat', 'none', '0']:
            return None
        try:
            if isinstance(val, (datetime, pd.Timestamp)):
                return val.strftime('%H:%M')
            if len(val_str) > 5:
                return val_str[:5]
            return val_str
        except:
            return None

    @classmethod
    def detect_header_row(cls, dosya_yolu, max_satir=30):
        """Akıllı Başlık Tespiti"""
        try:
            if dosya_yolu.endswith('.xls'):
                df_temp = pd.read_excel(dosya_yolu, header=None, nrows=max_satir, engine='xlrd')
            else:
                df_temp = pd.read_excel(dosya_yolu, header=None, nrows=max_satir, engine='openpyxl')
            
            KEYWORDS = ['TARIH', 'GIRIS', 'CIKIS', 'ADI', 'SOYADI', 'SAAT', 'SURE', 'GUN', 'SICIL', 'NORMAL']
            max_skor = 0
            en_iyi_satir = 0
            
            for index, row in df_temp.iterrows():
                satir_metni = " ".join([str(x).upper() for x in row if pd.notna(x)])
                # Başlık tespiti için de karakter düzeltme yapıyoruz
                satir_metni = satir_metni.replace('İ', 'I').replace('Ğ', 'G').replace('Ü', 'U').replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C')
                skor = sum(1 for k in KEYWORDS if k in satir_metni)
                if skor > max_skor:
                    max_skor = skor
                    en_iyi_satir = index
            
            return en_iyi_satir if max_skor >= 2 else 0
        except:
            return 0

    @classmethod
    def dosya_temizle(cls, dosya_yolu):
        try:
            header_index = cls.detect_header_row(dosya_yolu)
            
            if dosya_yolu.endswith('.xls'):
                df = pd.read_excel(dosya_yolu, header=header_index, engine='xlrd')
            else:
                df = pd.read_excel(dosya_yolu, header=header_index)

            # --- SÜTUN İSİMLERİNİ DÜZELT ---
            def sutun_duzelt(s):
                s = str(s).strip().upper()
                s = s.replace('CÝKÝÞ', 'CIKIS').replace('GIRIÞ', 'GIRIS')
                s = s.replace('Ý', 'I').replace('Þ', 'S').replace('Ð', 'G')
                s = s.replace('İ', 'I').replace('Ğ', 'G').replace('Ü', 'U')
                s = s.replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C')
                return s

            df.columns = [sutun_duzelt(col) for col in df.columns]

            giris_cols = [c for c in df.columns if 'GIRIS' in c]
            cikis_cols = [c for c in df.columns if 'CIKIS' in c]
            tarih_cols = [c for c in df.columns if 'TARIH' in c]
            ad_cols = [c for c in df.columns if 'ADI' in c or 'SOYADI' in c]
            gun_cols = [c for c in df.columns if c == 'GUN']
            pg_cols = [c for c in df.columns if 'PG' in c]

            if not giris_cols or not cikis_cols:
                return None

            tarih_col = tarih_cols[0]
            df = df.dropna(subset=[tarih_col])
            df['Temp_Date'] = pd.to_datetime(df[tarih_col], dayfirst=True, errors='coerce')
            df = df.dropna(subset=['Temp_Date'])

            # Final DataFrame
            clean_df = pd.DataFrame()
            
            # 1. AD SOYADI (Düzeltmeli)
            if ad_cols:
                clean_df['Adı Soyadı'] = df[ad_cols[0]].apply(cls.turkce_karakter_duzelt)
            else:
                clean_df['Adı Soyadı'] = "Personel"
            
            clean_df['Tarih'] = df['Temp_Date']
            
            # 2. GÜN (Düzeltmeli - Çarþamba -> Çarşamba)
            if gun_cols:
                clean_df['Gün'] = df[gun_cols[0]].apply(cls.turkce_karakter_duzelt)
            else:
                clean_df['Gün'] = ""
                
            # 3. PG / VARDİYA (Düzeltmeli)
            if pg_cols:
                clean_df['Pg.'] = df[pg_cols[0]].apply(cls.turkce_karakter_duzelt)
            else:
                clean_df['Pg.'] = ""
            
            # 4. Saatler (Formatlı)
            clean_df['Giriş'] = df[giris_cols[0]].apply(cls.saat_formatla)
            clean_df['Çıkış'] = df[cikis_cols[0]].apply(cls.saat_formatla)

            clean_df = clean_df[(clean_df['Giriş'].notna()) | (clean_df['Çıkış'].notna())]
            clean_df = clean_df.sort_values(by=['Tarih', 'Adı Soyadı'])
            
            return clean_df

        except Exception as e:
            st.error(f"Dosya işleme hatası ({os.path.basename(dosya_yolu)}): {e}")
            return None
    
    # calistir_etl metodu aynı kalıyor, yukarıdaki sınıfın içinde zaten mevcut.
    @classmethod
    def calistir_etl(cls, ham_klasor, hedef_klasor, hedef_dosya):
        dosyalar = glob.glob(os.path.join(ham_klasor, "*"))
        tum_veriler = []

        for dosya in dosyalar:
            if dosya.lower().endswith(('.xls', '.xlsx', '.csv')) and not os.path.basename(dosya).startswith('~$'):
                df = cls.dosya_temizle(dosya)
                if df is not None and len(df) > 0:
                    tum_veriler.append(df)

        if tum_veriler:
            ana_df = pd.concat(tum_veriler, ignore_index=True)
            ana_df = ana_df.sort_values(by=['Tarih', 'Adı Soyadı'])
            cikti_yolu = os.path.join(hedef_klasor, hedef_dosya)
            ana_df.to_excel(cikti_yolu, index=False)
            return cikti_yolu
        else:
            return None
# ============================================================================
# 4. BORDRO MOTORU
# ============================================================================
class PayrollEngine:
    """İş Kanunu hesaplamaları"""

    @staticmethod
    def mola_suresi_hesapla(brut_calisma_saati):
        if brut_calisma_saati <= 4:
            return 0.25
        elif brut_calisma_saati <= 7.5:
            return 0.50
        else:
            return 1.00

    @staticmethod
    def saat_farki_hesapla_net(giris, cikis):
        if pd.isna(giris) or pd.isna(cikis) or str(giris).strip() == '' or str(cikis).strip() == '':
            return 0.0
        try:
            fmt = "%H:%M" if len(str(giris)) <= 5 else "%H:%M:%S"
            g_dt = datetime.strptime(str(giris), fmt)
            c_dt = datetime.strptime(str(cikis), fmt)
            if c_dt < g_dt:
                c_dt += timedelta(days=1)
            fark = c_dt - g_dt
            brut_saat = fark.total_seconds() / 3600.0
            mola = PayrollEngine.mola_suresi_hesapla(brut_saat)
            return max(0.0, brut_saat - mola)
        except:
            return 0.0

    @staticmethod
    def ozel_yuvarlama(saat_val):
        if saat_val <= 0:
            return 0
        tam = int(saat_val)
        dakika = (saat_val - tam) * 60
        if dakika < 15:
            return float(tam)
        elif dakika < 45:
            return float(tam) + 0.5
        else:
            return float(tam) + 1.0

    @staticmethod
    def gece_calismasi_mi(giris, cikis):
        """
        Postalar Halinde İşçi Çalıştırılması Yönetmeliği Madde 7 uyarınca:
        Çalışma süresinin YARISINDAN ÇOĞU gece dönemine (20:00 - 06:00) 
        denk geliyorsa, bu çalışma 'Gece Çalışması' sayılır.
        """
        try:
            # Veri boşsa False dön
            if pd.isna(giris) or pd.isna(cikis):
                return False
            
            # String verileri saat objesine çevir
            str_giris = str(giris).strip()
            str_cikis = str(cikis).strip()
            if not str_giris or not str_cikis:
                return False

            fmt = "%H:%M" if len(str_giris) <= 5 else "%H:%M:%S"
            g_dt = datetime.strptime(str_giris[:5], "%H:%M")
            c_dt = datetime.strptime(str_cikis[:5], "%H:%M")

            # Gece yarısı geçişini yönet (Örn: 22:00 giriş, 02:00 çıkış)
            if c_dt < g_dt:
                c_dt += timedelta(days=1)

            # 1. Toplam Çalışma Süresini Bul (Saniye cinsinden)
            total_seconds = (c_dt - g_dt).total_seconds()
            
            if total_seconds <= 0:
                return False

            # 2. İlgili Gece Dönemini Belirle (20:00 - 06:00)
            # Eğer vardiya öğlen 12'den sonra başladıysa, bugünün gecesi baz alınır.
            # Eğer vardiya sabah 04:00 gibi başladıysa, bir önceki günün gecesi baz alınır.
            if g_dt.hour >= 12:
                night_start = g_dt.replace(hour=20, minute=0, second=0)
            else:
                night_start = (g_dt - timedelta(days=1)).replace(hour=20, minute=0, second=0)
            
            night_end = night_start + timedelta(hours=10) # 20:00 + 10 saat = 06:00

            # 3. Kesişimi (Overlap) Hesapla
            latest_start = max(g_dt, night_start)
            earliest_end = min(c_dt, night_end)
            
            overlap = (earliest_end - latest_start).total_seconds()
            overlap = max(0.0, overlap) # Negatif çıkarsa 0 yap

            # Toleransı saniyeye çevir ve düş
            tolerans_saniye = Config.GECE_CALISMA_TOLERANS_DK * 60
            efektif_gece_suresi = max(0.0, overlap - tolerans_saniye)
            # 4. KARAR ANI: Gece süresi, toplam sürenin yarısından fazla mı?
            return efektif_gece_suresi > (total_seconds / 2)

        except Exception as e:
            # Beklenmedik bir hata olursa güvenli tarafta kalıp False dönelim
            # print(f"Hata: {e}") 
            return False

    @staticmethod
    def decimal_hesapla(deger1, deger2=None, deger3=None, islem="carpma"):
        try:
            d1 = Decimal(str(deger1))
            if islem == "carpma":
                if deger2 is not None and deger3 is not None:
                    d2 = Decimal(str(deger2))
                    d3 = Decimal(str(deger3))
                    sonuc = d1 * d2 * d3
                elif deger2 is not None:
                    d2 = Decimal(str(deger2))
                    sonuc = d1 * d2
                else:
                    sonuc = d1
            elif islem == "cikarma":
                d2 = Decimal(str(deger2))
                sonuc = d1 - d2
            else:
                sonuc = d1
            return float(sonuc.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        except:
            return 0.0


# ============================================================================
# 5. BORDRO OKUYUCU (GÜNCELLENMİŞ - İSİM FİLTRELİ)
# ============================================================================
class BordroReader:

    @staticmethod
    def metni_sayiya_cevir(metin):
        if not metin or metin == '0' or metin == '0,00':
            return 0.0
        try:
            temiz = str(metin).replace('TL', '').replace('₺', '').strip()
            temiz = temiz.replace('.', '').replace(',', '.')
            return float(temiz)
        except:
            return 0.0

    @staticmethod
    def normalize_str(text):
        """
        Karşılaştırma için metni 'süper' temizler.
        1. Büyük harfe çevirir.
        2. Türkçe karakterleri İngilizce karşılıklarına dönüştürür.
        3. Harf ve Rakam DIŞINDAKİ her şeyi (boşluk, _, -, . vb.) siler.
        
        Örnek: "ECEVİT_ŞENGÜN" -> "ECEVITSENGUN"
               "Ecevit ŞENGÜN" -> "ECEVITSENGUN"
        Sonuç: Eşleşme Başarılı ✅
        """
        if not text: return ""
        
        # 1. Stringe çevir ve büyüt
        text = str(text).upper()
        
        # 2. Türkçe karakter dönüşüm haritası
        tr_map = {'İ': 'I', 'Ğ': 'G', 'Ü': 'U', 'Ş': 'S', 'Ö': 'O', 'Ç': 'C'}
        for tr, en in tr_map.items():
            text = text.replace(tr, en)
            
        # 3. Agresif Temizlik: Sadece A-Z ve 0-9 kalsın. 
        # (Boşluk, alt çizgi, nokta, virgül hepsi silinir)
        import re
        text = re.sub(r'[^A-Z0-9]', '', text)
        
        return text

    @staticmethod
    # BURAYA DİKKAT: Yeni parametre 'hedef_isim' eklendi varsayılan None
    def pdf_oku(pdf_path, hedef_isim=None):
        try:
            import pdfplumber
        except ImportError:
            st.error("⚠️ pdfplumber kurulu değil.")
            return BordroReader._manuel_veri()

        print("\n" + "="*60)
        print(f"📄 BORDRO OKUMA BAŞLIYOR... (Hedef Personel: {hedef_isim if hedef_isim else 'Filtresiz'})")
        print("="*60)

        if not os.path.exists(pdf_path):
            st.warning(f"⚠️ PDF bulunamadı: {pdf_path}")
            return BordroReader._manuel_veri()

        veriler = []
        PARA_REGEX = r'(?:\d{1,3}(?:\.\d{3})*|\d+)(?:,\d{1,2})?'
        
        # Kesme kelimeleri (değişmedi)
        KESME_KELIMELERI = [
            "SSK", "ssk", "İnd", "ind", "Bes ", "bes ", "Nafaka", "İcra", 
            "YASAL", "KESİNTİLER", "ÖZEL KES", "Küm", "Kum", "KÜM", "Matrah", 
            "MATRAH", "Gelir Verg", "G.Vergisi", "Damga", "Ters Bak"
        ]

        try:
            with pdfplumber.open(pdf_path) as pdf:
                for sayfa_no, page in enumerate(pdf.pages, 1):
                    text = page.extract_text()
                    if not text:
                        continue
                    
                    # FİLTRELEME MANTIĞI
                    if hedef_isim and hedef_isim.lower() != "personel":
                        norm_text = BordroReader.normalize_str(text)
                        norm_hedef = BordroReader.normalize_str(hedef_isim)
                        
                        # Basit kontrol: Hedef isim metin içinde geçiyor mu?
                        if norm_hedef not in norm_text:
                            # Debug Logu: Neden atlandığını göster
                            print(f"❌ Sayfa {sayfa_no} ATLANDI. (Aranan: '{hedef_isim}', Sayfada Eşleşme Bulunamadı)")
                            continue # Bir sonraki sayfaya geç
                        else:
                            print(f"✅ Sayfa {sayfa_no} İŞLENİYOR. (Personel Doğrulandı: '{hedef_isim}')")
                    # ---------------------------------------------------------

                    lines = text.split('\n')
                    current_donem = None
                    current_brut_ucret = 0.0
                    current_saat_ucreti = 0.0
                    current_gun_sayisi = 0
                    
                    # Değişkenleri sıfırla
                    p_fm, p_ubgt, p_ht, p_sorumluluk = 0.0, 0.0, 0.0, 0.0
                    p_ayni, p_izin, p_diger, p_agi, p_net = 0.0, 0.0, 0.0, 0.0, 0.0

                    # --- TARAMA MANTIĞI (Aynı kaldı) ---
                    for line in lines:
                        if "201" in line or "202" in line:
                            match = re.search(r'([a-zA-ZçÇğĞıİöÖşŞüÜ]+)\s+(20\d{2})', line)
                            if match:
                                ay_adi = match.group(1).title()
                                yil = match.group(2)
                                for k, v in Config.AYLAR.items():
                                    if k.lower() in ay_adi.lower():
                                        current_donem = f"{yil}-{v:02d}"
                                        break
                        if current_donem: break

                    print(f"   ---> İşlenen Dönem: {current_donem}")
                    
                    # 2. TARAMA: KALEMLER (Aynı kaldı)
                    for line in lines:
                        line_lower = line.lower()
                        if not re.search(r'\d+[,\.]\d{1,2}', line): continue

                        temiz_satir = line
                        if "fiili çalışma" not in line_lower and "normal çalışma" not in line_lower:
                            for kelime in KESME_KELIMELERI:
                                if kelime.lower() in temiz_satir.lower():
                                    idx = temiz_satir.lower().find(kelime.lower())
                                    if idx > 2: temiz_satir = temiz_satir[:idx]

                        raw_nums = re.findall(PARA_REGEX, temiz_satir)
                        if not raw_nums: continue

                        float_vals = [BordroReader.metni_sayiya_cevir(x) for x in raw_nums]
                        valid_vals = [x for x in float_vals if x > 0]

                        if not valid_vals: continue

                        val_tutar = max(valid_vals) 
                        val_miktar = 0
                        possible_amounts = [x for x in valid_vals if x < val_tutar and x < 400]
                        
                        if possible_amounts:
                            if "fiili çalışma" in line_lower:
                                val_miktar = possible_amounts[0]
                            else:
                                val_miktar = possible_amounts[-1]

                        # --- A. BRÜT ÜCRET HESABI ---
                        if "fiili çalışma" in line_lower or "normal çalışma" in line_lower or "normal kazanç" in line_lower:
                            if "fazla" not in line_lower:
                                if val_miktar == 0:
                                    basit_sayilar = re.findall(r'\b\d{1,2}[.,]\d{1,2}\b|\b\d{1,2}\b', temiz_satir)
                                    for s in basit_sayilar:
                                        s_float = BordroReader.metni_sayiya_cevir(s)
                                        if 0 < s_float < 32 and s_float < val_tutar:
                                            val_miktar = s_float
                                            break

                                if val_miktar > 0 and val_tutar > 0:
                                    current_gun_sayisi = val_miktar
                                    gunluk_ucret = val_tutar / val_miktar
                                    current_brut_ucret = gunluk_ucret * 30
                                    current_saat_ucreti = current_brut_ucret / 225
                                    
                                    if not (1 < current_saat_ucreti < 5000):
                                        current_brut_ucret = 0.0
                                        current_saat_ucreti = 0.0

                        # DİĞER KALEMLER
                        elif any(x in line_lower for x in ["fazla mesai", "f.mesai", "fm ", "f. mesai"]):
                            p_fm += val_tutar
                            if current_saat_ucreti == 0 and val_miktar > 0:
                                calculated = val_tutar / (val_miktar * 1.5)
                                if 50 < calculated < 2000:
                                    current_saat_ucreti = calculated
                                    current_brut_ucret = calculated * 225
                        
                        elif any(x in line_lower for x in ["genel tatil", "bayram mesai", "resmi tatil"]):
                            p_ubgt += val_tutar
                        elif any(x in line_lower for x in ["pazar mesai", "p.mesai"]):
                            p_ht += val_tutar
                        elif "sorumluluk" in line_lower: p_sorumluluk += val_tutar
                        elif "ayni yardım" in line_lower or "ayni yardim" in line_lower: p_ayni += val_tutar
                        elif "yıllık izin" in line_lower: p_izin += val_tutar
                        elif "diğer gelir" in line_lower or "diğer ek" in line_lower: p_diger += val_tutar
                        elif "agi " in line_lower or "asgari geçim" in line_lower: p_agi += val_tutar

                        if "toplam net" in line_lower or "net ödenen" in line_lower:
                             raw_net = re.findall(PARA_REGEX, line)
                             if raw_net:
                                 vals_net = [BordroReader.metni_sayiya_cevir(x) for x in raw_net]
                                 if vals_net: p_net = vals_net[-1]

                    # --- VERİ EKLEME (Aynı kaldı) ---
                    if current_donem:
                        mevcut = next((item for item in veriler if item["Donem_Kodu"] == current_donem), None)
                        if mevcut:
                            mevcut['Odenen_FM_TL'] += p_fm
                            mevcut['Odenen_UBGT_TL'] += p_ubgt
                            mevcut['Odenen_HT_TL'] += p_ht
                            mevcut['Odenen_Sorumluluk_TL'] += p_sorumluluk
                            mevcut['Odenen_Ayni_Yardim_TL'] += p_ayni
                            mevcut['Odenen_Yillik_Izin_TL'] += p_izin
                            mevcut['Odenen_Diger_TL'] += p_diger
                            mevcut['Odenen_AGI_TL'] += p_agi
                            if p_net > 0: mevcut['Odenen_Net_TL'] = p_net
                            if current_brut_ucret > 0:
                                mevcut['Bordro_Brut_Ucret'] = current_brut_ucret
                                mevcut['Bordro_Saat_Ucreti'] = current_saat_ucreti
                                mevcut['Calisilan_Gun_Sayisi'] = current_gun_sayisi
                        else:
                            veriler.append({
                                'Donem_Kodu': current_donem,
                                'Bordro_Brut_Ucret': round(current_brut_ucret, 2),
                                'Bordro_Saat_Ucreti': round(current_saat_ucreti, 2),
                                'Calisilan_Gun_Sayisi': current_gun_sayisi,
                                'Odenen_FM_TL': round(p_fm, 2),
                                'Odenen_UBGT_TL': round(p_ubgt, 2),
                                'Odenen_HT_TL': round(p_ht, 2),
                                'Odenen_Sorumluluk_TL': round(p_sorumluluk, 2),
                                'Odenen_Ayni_Yardim_TL': round(p_ayni, 2),
                                'Odenen_Yillik_Izin_TL': round(p_izin, 2),
                                'Odenen_Diger_TL': round(p_diger, 2),
                                'Odenen_AGI_TL': round(p_agi, 2),
                                'Odenen_Net_TL': round(p_net, 2)
                            })
                            
        except Exception as e:
            st.error(f"Hata: {e}")
            return BordroReader._manuel_veri()
            
        if veriler:
            df_ret = pd.DataFrame(veriler)
            # Eksik kolonları doldurma
            cols = ['Bordro_Brut_Ucret', 'Calisilan_Gun_Sayisi', 'Odenen_Sorumluluk_TL', 
                   'Odenen_Ayni_Yardim_TL', 'Odenen_Yillik_Izin_TL', 'Odenen_Diger_TL', 
                   'Odenen_AGI_TL', 'Odenen_Net_TL']
            for c in cols:
                if c not in df_ret.columns: df_ret[c] = 0.0
            
            return df_ret
            
        return BordroReader._manuel_veri()
    
    @staticmethod
    def _manuel_veri():
        # Burası aynı kalacak
        return pd.DataFrame([{
            'Donem_Kodu': '2024-01',
            'Bordro_Brut_Ucret': 0.0,
            'Bordro_Saat_Ucreti': 0.0,
            'Calisilan_Gun_Sayisi': 0,
            'Odenen_FM_TL': 0.0,
            'Odenen_UBGT_TL': 0.0,
            'Odenen_HT_TL': 0.0,
            'Odenen_Net_TL': 0.0
        }])
    
# ============================================================================
# 6. EXCEL ÜRETICI
# ============================================================================
class ExcelGenerator:
    @staticmethod
    def gorsel_puantaj_olustur(temiz_veri_yolu, cikti_klasoru):

        print("\n" + "="*60)
        print("📊 GÖRSEL PUANTAJ (VERİ ÜRETİCİ)")
        print("="*60)

        df = pd.read_excel(temiz_veri_yolu)
        df['Tarih'] = pd.to_datetime(df['Tarih'])
        df['Hafta_Basi'] = df['Tarih'].apply(lambda x: x - timedelta(days=x.weekday()))
        
        # Agresif Karakter Temizleme
        def temizle_metin(s):
            s = str(s).strip().upper()
            s = s.replace('Ý', 'I').replace('Þ', 'S').replace('Ð', 'G')
            s = s.replace('İ', 'I').replace('Ğ', 'G').replace('Ü', 'U').replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C')
            return s
        
        df['Gün'] = df['Gün'].apply(temizle_metin)
        df['Pg.'] = df['Pg.'].apply(temizle_metin)
            
        # --- TEK VE MERKEZİ HESAPLAMA ---
        def analiz_et(row):
            tarih = row['Tarih'].date()
            pg = str(row.get('Pg.', '')).upper()
            gun = str(row.get('Gün', '')).upper()
            
            calisti = (str(row['Giriş']) not in ['nan', '', 'NaT'] and
                      str(row['Çıkış']) not in ['nan', '', 'NaT'])
            is_gece = False
            if calisti:
                is_gece = PayrollEngine.gece_calismasi_mi(row['Giriş'], row['Çıkış'])
            
            ham = PayrollEngine.saat_farki_hesapla_net(row['Giriş'], row['Çıkış'])
            puan = PayrollEngine.ozel_yuvarlama(ham)
            
            if Config.MINIMUM_SURE_GARANTISI and calisti and 0 < puan < 7.5:
                puan = 7.5
                
            durum = "NORMAL"
            if "HAFTA TATILI" in pg or gun == 'PAZAR':
                durum = "HT"
            elif StaticHolidays.is_arefe(tarih):
                durum = "AREFE"
            elif StaticHolidays.is_ubgt(tarih) or "GENEL TATIL" in pg:
                durum = "UBGT"
            
            val = puan if calisti else Config.HAFTA_TATILI_ISARETI
            # Final_Hesap: Sayısal hesaplamalar için kullanılacak net saat
            final_hesap = puan if calisti else 0.0
            
            return pd.Series([durum, val, final_hesap, is_gece])

        df[['Durum', 'Puan_Degeri', 'Final_Hesap', 'Gece_Mi']] = df.apply(analiz_et, axis=1)

        # --- EXCEL GÖRSELLEŞTİRME (PIVOT) ---
        pivot = df.pivot_table(index='Hafta_Basi', columns='Gün', values='Puan_Degeri', aggfunc='first')
        
        if not df.empty:
            min_hb = df['Hafta_Basi'].min()
            max_hb = df['Hafta_Basi'].max()
            full_weeks = pd.date_range(start=min_hb, end=max_hb, freq='7D')
            pivot = pivot.reindex(full_weeks)
        
        temiz_gun_isimleri = [
            g.upper()
            .replace('İ', 'I').replace('Ğ', 'G').replace('Ü', 'U')
            .replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C') 
            for g in Config.GUN_ISIMLERI
        ]
        
        pivot = pivot.reindex(columns=temiz_gun_isimleri, fill_value="x").fillna("x")
        pivot.index.name = 'Hafta_Basi'
        pivot = pivot.reset_index()
        pivot['Hafta_Sonu'] = pivot['Hafta_Basi'] + timedelta(days=6)

        # Görsel Tablo İçin Haftalık Hesaplama (Sadece Excel'e basmak için)
        def hesapla_haftalik_gorsel(row):
            h_basi = row['Hafta_Basi']
            veri = df[df['Hafta_Basi'] == h_basi]
            calisilan = veri[veri['Final_Hesap'] > 0]
            toplam = calisilan['Final_Hesap'].sum()
            ubgt_gun = 0
            ht_gun = 0
            dusulecek = 0
            for _, r in veri.iterrows():
                if r['Final_Hesap'] > 0:
                    if r['Durum'] == 'HT':
                        ht_gun += 1
                        dusulecek += min(r['Final_Hesap'], 7.5)
                    elif r['Durum'] == 'UBGT':
                        ubgt_gun += 1
                        dusulecek += min(r['Final_Hesap'], 7.5)
                    elif r['Durum'] == 'AREFE':
                        ubgt_gun += 0.5
                        dusulecek += min(r['Final_Hesap'], 7.5)
            fm = 0
            if calisilan['Gece_Mi'].any():
                for s in calisilan['Final_Hesap']:
                    fm += max(0, s - 7.5)
            else:
                mesaiye_esas = toplam - dusulecek
                fm = max(0, mesaiye_esas - Config.HAFTALIK_YASAL_SURE)
            return pd.Series([toplam, fm, ubgt_gun, ht_gun])

        pivot[['Haftalık Ç.S.', 'FM Saati', 'UBGT', 'HT']] = pivot.apply(hesapla_haftalik_gorsel, axis=1)
        
        # Excel Kaydetme İşlemleri
        pivot['FM Saati'] = pivot['FM Saati'].replace(0, '')
        pivot['UBGT'] = pivot['UBGT'].replace(0, '')
        pivot['HT'] = pivot['HT'].replace(0, '')
        pivot['D_Bas'] = pivot['Hafta_Basi'].dt.strftime('%d.%m.%Y')
        pivot['D_Bit'] = pivot['Hafta_Sonu'].dt.strftime('%d.%m.%Y')

        final_cols = ['D_Bas', 'D_Bit'] + temiz_gun_isimleri + ['Haftalık Ç.S.', 'FM Saati', 'UBGT', 'HT']

        ad = df['Adı Soyadı'].dropna().iloc[0] if not df['Adı Soyadı'].dropna().empty else "Personel"
        temiz_ad = "".join([c if c.isalnum() else "_" for c in str(ad)]).strip()
        cikti_adi = f"{temiz_ad}_PUANTAJ.xlsx"
        cikti_yolu = os.path.join(cikti_klasoru, cikti_adi)

        writer = pd.ExcelWriter(cikti_yolu, engine='openpyxl')
        pivot[final_cols].to_excel(writer, sheet_name='PUANTAJ', startrow=3, index=False, header=False)
        ws = writer.book['PUANTAJ']
        
        # --- Stil İşlemleri ---
        sari = PatternFill("solid", fgColor=Config.RENK_SARI)
        gri = PatternFill("solid", fgColor=Config.RENK_GRI)
        yesil = PatternFill("solid", fgColor=Config.RENK_YESIL)
        sari_ubgt = PatternFill("solid", fgColor=Config.RENK_FOSFOR)
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        ws['A1'] = "4857 SAYILI İŞ KANUNU PUANTAJ"
        ws['A1'].fill = sari
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:M1')
        ws['A1'].alignment = Alignment("center", "center")
        ws['A3'] = "DÖNEM"
        ws['A3'].fill = gri

        basliklar = {'C3': 'Pzt', 'D3': 'Sal', 'E3': 'Çar', 'F3': 'Per', 'G3': 'Cum', 'H3': 'Cmt', 'I3': 'Paz', 'J3': 'Top. Saat', 'K3': 'FM', 'L3': 'UBGT', 'M3': 'HT'}
        for k, v in basliklar.items():
            ws[k].value = v
            ws[k].fill = gri
            ws[k].border = border
            ws[k].alignment = Alignment("center", "center")

        data_rows = ws[f'A4:M{ws.max_row}']
        gun_map = {2: 0, 3: 1, 4: 2, 5: 3, 6: 4, 7: 5, 8: 6}

        for r_idx, row in enumerate(data_rows):
            h_basi = pivot.iloc[r_idx]['Hafta_Basi']
            for cell in row:
                cell.border = border
                cell.alignment = Alignment("center", "center")
                c_idx = cell.col_idx - 1
                if c_idx in gun_map:
                    tarih = h_basi + timedelta(days=gun_map[c_idx])
                    kayit = df[df['Tarih'] == tarih]
                    if not kayit.empty:
                        durum = kayit.iloc[0]['Durum']
                        if durum == 'AREFE':
                            cell.fill = yesil
                        elif durum == 'UBGT':
                            cell.fill = sari_ubgt
                        if kayit.iloc[0]['Gece_Mi']:
                            cell.font = Font(bold=True, color=Config.RENK_KIRMIZI)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        writer.close()
        
        print(f"✅ Puantaj ve Veri Hazır: {cikti_adi}")
        
        # KRİTİK: Hem dosya yolunu hem de işlenmiş DataFrame'i döndür
        return cikti_yolu, df, temiz_ad

    @staticmethod
    def alacak_raporu_olustur(processed_df, bordro_df, cikti_klasoru, dosya_oneki=""):
        print("\n" + "="*60)
        print("💰 HASSAS ALACAK RAPORU (PUANTAJ VERİSİ İLE)")
        print("="*60)
    
        # Dosya ismine ön ek ekle (Örn: ECEVIT_SENGUN_HASSAS_ALACAK_RAPORU.xlsx)
        if dosya_oneki:
            dosya_adi = f"{dosya_oneki}_ALACAK_RAPORU.xlsx"
        else:
            dosya_adi = "HASSAS_ALACAK_RAPORU.xlsx"
        cikti_yolu = os.path.join(cikti_klasoru, dosya_adi)
        
        # 1. Puantajdan gelen veriyi kopyala
        df = processed_df.copy()
        
        # 2. Haftalık ve Günlük Dönemleri Ayarla
        df['Hafta_Sonu'] = df['Hafta_Basi'] + timedelta(days=6) # Pazar Günü
        
        # FM ve HT -> Haftanın bittiği ayın bordrosuna
        df['Donem_Haftalik'] = df['Hafta_Sonu'].dt.strftime('%Y-%m')
        
        # UBGT -> Olayın olduğu ayın bordrosuna
        df['Donem_Gunluk'] = df['Tarih'].dt.strftime('%Y-%m')

        # --- HAFTALIK HESAPLAMA MOTORU ---
        haftalik_sonuclar = []

        for hb, grup in df.groupby('Hafta_Basi'):
            donem_kodu = grup['Hafta_Sonu'].iloc[0].strftime('%Y-%m')
            
            calisan = grup[grup['Final_Hesap'] > 0]
            toplam_saat = calisan['Final_Hesap'].sum()
            
            ht_gun_sayisi = len(grup[(grup['Durum'] == 'HT') & (grup['Final_Hesap'] > 0)])
            
            dusulecek = 0
            for _, r in grup.iterrows():
                if r['Final_Hesap'] > 0 and r['Durum'] in ['HT', 'UBGT', 'AREFE']:
                    dusulecek += min(r['Final_Hesap'], 7.5)
            
            gece_fm = 0
            if calisan['Gece_Mi'].any():
                for s in calisan['Final_Hesap']:
                    if s > 7.5: gece_fm += (s - 7.5)
            
            normal_fm = max(0, (toplam_saat - dusulecek) - Config.HAFTALIK_YASAL_SURE)
            haftalik_fm = max(gece_fm, normal_fm)
            
            haftalik_sonuclar.append({
                'Donem_Kodu': donem_kodu,
                'Hak_FM_Saat': haftalik_fm,
                'Hak_HT_Gun': ht_gun_sayisi
            })

        df_haftalik = pd.DataFrame(haftalik_sonuclar)
        ozet_fm_ht = df_haftalik.groupby('Donem_Kodu').sum().reset_index()

        # --- UBGT HESAPLAMA ---
        def ubgt_skor(row):
            if row['Final_Hesap'] > 0:
                if row['Durum'] == 'UBGT': return 1.0
                if row['Durum'] == 'AREFE': return 0.5
            return 0.0
        
        df['UBGT_Skor'] = df.apply(ubgt_skor, axis=1)
        ozet_ubgt = df.groupby('Donem_Gunluk')['UBGT_Skor'].sum().reset_index()
        ozet_ubgt.rename(columns={'Donem_Gunluk': 'Donem_Kodu', 'UBGT_Skor': 'Hak_UBGT_Gun'}, inplace=True)

        # --- BİRLEŞTİRME VE MERGE ---
        ozet = pd.merge(ozet_fm_ht, ozet_ubgt, on='Donem_Kodu', how='outer').fillna(0)
        
        # Bordro verileriyle birleştir (how='outer' yaparak her iki taraftaki veriyi de koruyoruz)
        ana = pd.merge(ozet, bordro_df, on='Donem_Kodu', how='outer').fillna(0)
        ana = ana.sort_values('Donem_Kodu') # Kronolojik sıralama burada yapılıyor
        
        # --- TARİH FORMATI DEĞİŞİKLİĞİ (Haz.17) ---
        # Sıralama bittikten sonra görsel formatı uyguluyoruz
        def format_donem_tr(donem_str):
            try:
                # Gelen format: "2017-06"
                yil, ay = str(donem_str).split('-')
                ay_map = {
                    '01': 'Oca', '02': 'Şub', '03': 'Mar', '04': 'Nis', '05': 'May', '06': 'Haz',
                    '07': 'Tem', '08': 'Ağu', '09': 'Eyl', '10': 'Eki', '11': 'Kas', '12': 'Ara'
                }
                kisa_yil = yil[2:] # 2017 -> 17
                kisa_ay = ay_map.get(ay, '???')
                return f"{kisa_ay}.{kisa_yil}"
            except:
                return donem_str

        # Orijinal 'Donem_Kodu'nu formatlıyoruz
        ana['Donem_Kodu'] = ana['Donem_Kodu'].apply(format_donem_tr)
        
        # --- FİNANSAL HESAPLAMALAR ---
        ana['Bordro_Saat_Ucreti'] = ana['Bordro_Saat_Ucreti'].replace(0, np.nan).ffill().bfill().fillna(0).astype(float).round(2)
    
        ana['Hak_FM_Saat'] = ana['Hak_FM_Saat'].apply(lambda x: float(Decimal(str(x)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)))
        
        for col in ['Hak_FM_TL', 'Hak_UBGT_TL', 'Hak_HT_TL', 'Fark_FM_TL', 'Fark_UBGT_TL', 'Fark_HT_TL']:
            ana[col] = 0.0
    
        for idx, row in ana.iterrows():
            saat_ucreti = row['Bordro_Saat_Ucreti']
            gunluk_ucret = PayrollEngine.decimal_hesapla(saat_ucreti, 7.5)
            
            hak_fm_tl = PayrollEngine.decimal_hesapla(row['Hak_FM_Saat'], saat_ucreti, 1.5)
            hak_ubgt_tl = PayrollEngine.decimal_hesapla(row['Hak_UBGT_Gun'], gunluk_ucret)
            hak_ht_tl = PayrollEngine.decimal_hesapla(row['Hak_HT_Gun'], gunluk_ucret, 1.5)
    
            ana.at[idx, 'Hak_FM_TL'] = hak_fm_tl
            ana.at[idx, 'Hak_UBGT_TL'] = hak_ubgt_tl
            ana.at[idx, 'Hak_HT_TL'] = hak_ht_tl
            ana.at[idx, 'Fark_FM_TL'] = max(0, PayrollEngine.decimal_hesapla(hak_fm_tl, row['Odenen_FM_TL'], islem="cikarma"))
            ana.at[idx, 'Fark_UBGT_TL'] = max(0, PayrollEngine.decimal_hesapla(hak_ubgt_tl, row['Odenen_UBGT_TL'], islem="cikarma"))
            ana.at[idx, 'Fark_HT_TL'] = max(0, PayrollEngine.decimal_hesapla(hak_ht_tl, row['Odenen_HT_TL'], islem="cikarma"))
    
        # EXCEL YAZMA
        writer = pd.ExcelWriter(cikti_yolu, engine='openpyxl')
    
        sutunlar = {
            'Fazla Mesai': ['Donem_Kodu', 'Bordro_Saat_Ucreti', 'Hak_FM_Saat', 'Hak_FM_TL', 'Odenen_FM_TL', 'Fark_FM_TL'],
            'UBGT': ['Donem_Kodu', 'Bordro_Saat_Ucreti', 'Hak_UBGT_Gun', 'Hak_UBGT_TL', 'Odenen_UBGT_TL', 'Fark_UBGT_TL'],
            'Hafta Tatili': ['Donem_Kodu', 'Bordro_Saat_Ucreti', 'Hak_HT_Gun', 'Hak_HT_TL', 'Odenen_HT_TL', 'Fark_HT_TL']
        }
    
        toplam_alacak = 0
    
        def sheet_yap(isim, veri, renk):
            # Veriyi önce yuvarla (Matematiksel garanti)
            if not veri.empty:
                cols_to_round = veri.columns[1:] 
                veri[cols_to_round] = veri[cols_to_round].astype(float).round(2)

            veri.to_excel(writer, sheet_name=isim, index=False, startrow=2)
            ws = writer.book[isim]
            ws['A1'] = f"{isim} FARK CETVELİ"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill("solid", fgColor=renk)
            ws.merge_cells('A1:F1')
            ws['A1'].alignment = Alignment("center", "center")
            
            t = veri.iloc[:, -1].sum()
            lr = ws.max_row + 2
            ws[f'A{lr}'] = "TOPLAM:"
            ws[f'F{lr}'] = t
            ws[f'F{lr}'].number_format = '#,##0.00 TL'
            ws[f'F{lr}'].font = Font(bold=True, color=Config.RENK_KIRMIZI, size=12)
            
            # Stil döngüsü
            for row in ws[f'F3:F{ws.max_row}']:
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.fill = PatternFill("solid", fgColor="FFE6E6")
                        cell.font = Font(bold=True, color=Config.RENK_KIRMIZI)
            return t
    
        t1 = ana[sutunlar['Fazla Mesai']].copy()
        t1.columns = ['Dönem', 'Saat Ücreti', 'Hesaplanan (Saat)', 'Hesaplanan TL', 'Bordro Ödenen', 'FARK']

        toplam_alacak += sheet_yap('FAZLA_MESAI', t1, Config.RENK_MAVI)
    
        t2 = ana[sutunlar['UBGT']].copy()
        t2.columns = ['Dönem', 'Saat Ücreti', 'Hesaplanan (Gün)', 'Hesaplanan TL', 'Bordro Ödenen', 'FARK']
        t2 = t2[t2['FARK'] > 0]
        toplam_alacak += sheet_yap('UBGT', t2, Config.RENK_KOYU_KIRMIZI)
    
        t3 = ana[sutunlar['Hafta Tatili']].copy()
        t3.columns = ['Dönem', 'Saat Ücreti', 'Hesaplanan (Gün)', 'Hesaplanan TL', 'Bordro Ödenen', 'FARK']
        t3 = t3[t3['FARK'] > 0]
        toplam_alacak += sheet_yap('HAFTA_TATILI', t3, Config.RENK_KOYU_MAVI)
    
        ozet_df = pd.DataFrame([['GENEL TOPLAM ALACAK', toplam_alacak]], columns=['Kalem', 'Tutar'])
        ozet_df.to_excel(writer, sheet_name='OZET', index=False)
        writer.book['OZET']['B2'].number_format = '#,##0.00 TL'
        writer.book['OZET']['B2'].font = Font(bold=True, size=14, color=Config.RENK_KIRMIZI)
    
        # BORDRO SEKMESİ
        b_export = ana.copy()
        
        # Eğer Bordro_Brut_Ucret sütunu doluysa onu kullan, yoksa (puantajdan geldiyse) saatten hesapla
        if 'Bordro_Brut_Ucret' in b_export.columns:
            b_export['Aylik_Ucret_Brut'] = b_export['Bordro_Brut_Ucret'].fillna(0)
            # Eğer bordroda brüt 0 gelmişse (sadece puantaj varsa) saatten hesapla
            mask = b_export['Aylik_Ucret_Brut'] == 0
            b_export.loc[mask, 'Aylik_Ucret_Brut'] = round(b_export.loc[mask, 'Bordro_Saat_Ucreti'] * 225, 2)
        else:
            b_export['Aylik_Ucret_Brut'] = round(b_export['Bordro_Saat_Ucreti'] * 225, 2)

        # Günlük ücreti de Aylık / 30 mantığıyla göster (Sizin istediğiniz mantık)
        b_export['Gunluk_Ucret_Brut'] = round(b_export['Aylik_Ucret_Brut'] / 30, 2)

        b_export['FM_Saati'] = b_export['Hak_FM_Saat']
        
        # Eksik sütunları tamamla
        yeni_sutunlar = ['Odenen_Sorumluluk_TL', 'Odenen_Ayni_Yardim_TL', 'Odenen_Yillik_Izin_TL', 'Odenen_Diger_TL', 'Odenen_AGI_TL', 'Odenen_Net_TL']
        for c in yeni_sutunlar:
            if c not in b_export.columns: b_export[c] = 0.0
        if 'Banka_Odemesi' not in b_export.columns: b_export['Banka_Odemesi'] = 0.0
        b_export['Bordro_Imza'] = "yok"
        
        # SÜTUN İSİMLENDİRME (ORİJİNAL İSİMLER KORUNDU)
        rename_map = {
            'Donem_Kodu': 'DÖNEM',
            'Aylik_Ucret_Brut': 'AYLIK ÜCRET (Brüt)',
            'Gunluk_Ucret_Brut': 'GÜNLÜK ÜCRET (Brüt)',
            
            # Burada 'FM_Saati' artık bizim hesapladığımız veriyi taşıyor
            'FM_Saati': 'F.M. Saati', 
            'Odenen_FM_TL': 'F.M  Ücreti',
            
            # Hafta Tatili ve UBGT Günleri de bizim hesapladığımız verilerden gelsin
            'Hak_HT_Gun': 'H.T Günü',       # Bizim hesapladığımız gün sayısı
            'Odenen_HT_TL': ' H.T Ücreti',  # Bordroda ödenen para
            
            'Hak_UBGT_Gun': 'UBGT Günü',    # Bizim hesapladığımız gün sayısı
            'Odenen_UBGT_TL': ' UBGT Ücreti', # Bordroda ödenen para
            
            'Odenen_Diger_TL': 'Diğer Ek Ödeme ',
            'Odenen_Sorumluluk_TL': 'Sorumluluk Ücreti',
            'Odenen_Ayni_Yardim_TL': 'AYNI YARDIM',
            'Odenen_Yillik_Izin_TL': 'YILLIK İZİN ',
            'Odenen_AGI_TL': 'AGİ ',
            'Odenen_Net_TL': 'Bordro  Ödenen NET  Ücret',
            'Banka_Odemesi': 'Banka  Ödemesi',
            'Bordro_Imza': 'Bordro  İmza '
        }
        
        # Sütunları seç ve yeniden adlandır
        b_export = b_export.rename(columns=rename_map)
        
        # Sadece map'te tanımlı olan sütunları al
        final_cols = [col for col in rename_map.values() if col in b_export.columns]
        b_export = b_export[final_cols]
        
        # Sütun isimleri yukarıdaki rename_map'ten geliyor
        bosaltilacak_sutunlar = [
            'Diğer Ek Ödeme ', 
            'Sorumluluk Ücreti', 
            'AYNI YARDIM', 
            'YILLIK İZİN ', 
            'AGİ '
        ]
        
        for col in bosaltilacak_sutunlar:
            if col in b_export.columns:
                # 0 değerlerini NaN (boş) yapıyoruz. 
                # (Excel'de hücre boş görünür)
                b_export[col] = b_export[col].replace({0: np.nan, 0.0: np.nan})

        # Excel'e Yaz
        b_export.to_excel(writer, sheet_name='Bordro', index=False)
        
        # FORMATLAMA
        ws_b = writer.book['Bordro']
        header_fill = PatternFill("solid", fgColor="D9D9D9")
        
        for cell in ws_b[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        para_format = '#,##0.00 "₺"'
        for row in ws_b.iter_rows(min_row=2, max_col=len(final_cols)):
            for cell in row:
                # Kolon indekslerine göre formatlama (Basit mantık: Tutar içerenler para formatı)
                # Başlık ismini kontrol ederek format verelim
                col_name = ws_b.cell(1, cell.col_idx).value
                if col_name and any(x in col_name for x in ['Ücret', 'Tutar', 'Ödeme', 'YARDIM', 'AGİ']):
                     cell.number_format = para_format
                elif col_name and any(x in col_name for x in ['Saat', 'Gün']):
                     cell.number_format = '0.00'

        ws_b.column_dimensions['A'].width = 15
        ws_b.column_dimensions['B'].width = 20
        ws_b.column_dimensions['M'].width = 25
    
        writer.close()
        print(f"✅ Alacak raporu: {Config.ALACAK_RAPORU_DOSYASI}")
        return ana
    
# ============================================================================
# 7. ORKESTRATÖR
# ============================================================================
import streamlit as st
import os
import shutil
from datetime import datetime

# --- TASARIM AYARLARI ---
st.set_page_config(page_title="Puantaj Otomasyon Sistemi - <3", layout="wide")

# .tsx dosyanızdaki görsel dili yakalamak için özel CSS
st.markdown("""
    <style>
    .main { background-color: #f8f9fa; }
    .stButton>button { width: 100%; border-radius: 10px; height: 3em; background-color: #2563eb; color: white; }
    </style>
    """, unsafe_allow_html=True)

def main():
    # 1. HAFIZA (SESSION STATE) TANIMLARI
    if 'reset_counter' not in st.session_state:
        st.session_state.reset_counter = 0
    if 'hesaplama_tamam' not in st.session_state:
        st.session_state.hesaplama_tamam = False
    if 'metrikler' not in st.session_state:
        st.session_state.metrikler = {}  
    if 'ham_veri_onizleme' not in st.session_state:
        st.session_state.ham_veri_onizleme = None

    st.title("📊 Puantaj Otomasyon Sistemi - <3")
    st.info("Kullanıcıya özel izole çalışma alanı aktif. Umarım düzgün çalışır :)")

    paths = Config.get_paths()

    # --- SIDEBAR (AYNI KALDI) ---
    with st.sidebar:
        st.header("📁 Dosya Yükleme")
        excel_files = st.file_uploader(
            "Excel Dosyaları", 
            accept_multiple_files=True, 
            type=['xls', 'xlsx'], 
            key=f"ex_up_{st.session_state.reset_counter}"
        )
        pdf_file = st.file_uploader(
            "Bordro PDF", 
            type=['pdf'], 
            key=f"pdf_up_{st.session_state.reset_counter}"
        )

    st.divider()

    with st.sidebar.expander("⚙️ Hesaplama Parametreleri", expanded=False):
        st.caption("Varsayılan yasal değerleri buradan değiştirebilirsiniz.")
        yeni_haftalik_sure = st.number_input("Haftalık Yasal Süre (Saat)", min_value=1.0, max_value=60.0, value=45.0, step=0.5)
        yeni_gunluk_sure = st.number_input("Günlük Standart (Saat)", min_value=1.0, max_value=24.0, value=7.5, step=0.5)
        yeni_min_garanti = st.checkbox("Min. Süre Garantisi (7.5 Sa)", value=True)
        tolerans_dk = st.number_input("Gece Geçiş Toleransı (Dakika)", min_value=0, max_value=300, value=0)
        
        if st.sidebar.button("♻️ Tüm Verileri ve UI'ı Sıfırla"):
            if os.path.exists(paths["BASE"]):
                shutil.rmtree(paths["BASE"])
            st.session_state.hesaplama_tamam = False
            st.session_state.metrikler = {}
            st.session_state.ham_veri_onizleme = None # Veriyi de sıfırla
            st.session_state.reset_counter += 1
            st.rerun()

    # ============================================================================
    # AKIŞ KONTROLÜ (WIZARD MANTIĞI)
    # ============================================================================

    # DURUM 1: Henüz analiz yapılmamışsa "Analiz Et" butonunu göster
    # (Veri varsa veya Hesaplama bitmişse bu buton gizlenir)
    if st.session_state.ham_veri_onizleme is None and not st.session_state.hesaplama_tamam:
        
        analyze_btn = st.button("🔍 Verileri Analiz Et", use_container_width=True)

        if analyze_btn:
            if not excel_files:
                st.warning("⚠️ Lütfen önce Excel dosyalarını yükleyin.")
            else:
                paths = Config.klasorleri_hazirla()
                
                with st.status("Veriler taranıyor ve temizleniyor...", expanded=True) as status:
                    try:
                        for f in excel_files:
                            with open(os.path.join(paths["HAM"], f.name), "wb") as buffer:
                                shutil.copyfileobj(f, buffer)
                        
                        temiz_yol = ETLWorker.calistir_etl(paths["HAM"], paths["PDKS"], "Temiz_Veri.xlsx")
                        
                        if temiz_yol:
                            df_preview = pd.read_excel(temiz_yol)
                            df_preview['Tarih'] = pd.to_datetime(df_preview['Tarih'])
                            st.session_state.ham_veri_onizleme = df_preview
                            
                            status.update(label="✅ Analiz Tamamlandı!", state="complete", expanded=False)
                            time.sleep(0.5)
                            st.rerun() # EKRANI YENİLE (Butonu gizlemek için şart)
                        else:
                            status.update(label="❌ Veri okunamadı!", state="error")
                    except Exception as e:
                        status.update(label="❌ Hata", state="error")
                        st.error(str(e))

    # DURUM 2: Analiz yapılmış AMA henüz hesaplama bitmemişse -> Tabloyu göster
    elif st.session_state.ham_veri_onizleme is not None and not st.session_state.hesaplama_tamam:
        
        # 1. TÜM BU BÖLÜMÜ BİR 'PLACEHOLDER' İÇİNE ALIYORUZ
        preview_container = st.empty()
        
        # 2. Tablo ve Butonu bu kutunun içine çiziyoruz
        with preview_container.container():
            st.subheader("📝 Veri Önizleme ve Düzeltme")
            st.info("Tablodaki hatalı saatleri (Örn: 08:00) üzerine çift tıklayarak düzeltebilirsiniz.")

            edited_df = st.data_editor(
                st.session_state.ham_veri_onizleme,
                num_rows="dynamic",
                use_container_width=True,
                height=400,
                key="data_editor_key",
                column_config={
                    "Giriş": st.column_config.TextColumn("Giriş Saati", help="Format: 08:00"),
                    "Çıkış": st.column_config.TextColumn("Çıkış Saati", help="Format: 18:00"),
                    "Tarih": st.column_config.DateColumn("Tarih", format="YYYY-MM-DD"),
                }
            )

            st.write("---")
            
            # Buton da kutunun içinde
            hesapla_btn = st.button("🚀 Onayla ve Hesaplamayı Başlat", type="primary", use_container_width=True)
            
        if hesapla_btn:
            # Sihirli Satır: Butona basınca tabloyu ve butonu ekrandan siler
            preview_container.empty() 
            
            if pdf_file is None:
                st.error("⚠️ Lütfen sol menüden Bordro PDF dosyasını yükleyin!")
            else:
                paths = Config.get_paths()
                with open(paths["BORDRO_FILE"], "wb") as buffer:
                    shutil.copyfileobj(pdf_file, buffer)
                
                # Ayarları Enjekte Et
                Config.HAFTALIK_YASAL_SURE = yeni_haftalik_sure
                Config.GUNLUK_STANDART_SAAT = yeni_gunluk_sure
                Config.GECE_CALISMA_TOLERANS_DK = tolerans_dk
                Config.MINIMUM_SURE_GARANTISI = yeni_min_garanti
                
                with st.status("Final raporlar hazırlanıyor...", expanded=True) as status:
                    try:
                        # 1. Düzeltilmiş veriyi kaydet
                        revize_yol = os.path.join(paths["PDKS"], "Revize_Veri.xlsx")
                        edited_df.to_excel(revize_yol, index=False)
                        st.write("✅ Veriler işlendi.")

                        # --- 1. PUANTAJ HESAPLAMA (Dinamik Kontrol) ---
                        st.write("🔍 Puantaj hesaplanıyor...")
                        raw_puantaj = ExcelGenerator.gorsel_puantaj_olustur(revize_yol, paths["PUANTAJ"])
                        
                        # Gelen veri 3'lü mü (yol, df, isim) yoksa 2'li mi (yol, df)?
                        if isinstance(raw_puantaj, tuple) and len(raw_puantaj) == 3:
                            puantaj_yolu, processed_df, personel_adi = raw_puantaj
                        elif isinstance(raw_puantaj, tuple) and len(raw_puantaj) == 2:
                            puantaj_yolu, processed_df = raw_puantaj
                            personel_adi = "Personel"
                        else:
                            # Çok eski versiyon ise sadece yol dönüyor olabilir
                            puantaj_yolu = raw_puantaj
                            processed_df = pd.read_excel(puantaj_yolu) # Mecburen tekrar oku
                            personel_adi = "Personel"

                        st.write(f"📊 Puantaj Hazır: {personel_adi}")
                        
                        st.write("📄 Bordro PDF okunuyor...")
                        bordro_df = BordroReader.pdf_oku(paths["BORDRO_FILE"], hedef_isim=personel_adi)
                        
                        st.write("💰 Alacaklar hesaplanıyor...")
                        
                        # --- 2. ALACAK HESAPLAMA (HATA ÇIKAN YER BURASIYDI) ---
                        # Burayı "Akıllı Kontrol" ile sarıyoruz.
                        raw_alacak = ExcelGenerator.alacak_raporu_olustur(
                            processed_df, 
                            bordro_df, 
                            paths["RAPOR"], 
                            dosya_oneki=personel_adi 
                        )
                        
                        # GELEN VERİNİN TİPİNE BAKIP ONA GÖRE DAVRANIYORUZ:
                        if isinstance(raw_alacak, tuple) or isinstance(raw_alacak, list):
                            # Yeni versiyon: (DosyaYolu, DataFrame)
                            alacak_sonucu_path = raw_alacak[0]
                            alacak_sonucu_df = raw_alacak[1]
                        else:
                            # Eski versiyon: Sadece DataFrame dönmüş
                            alacak_sonucu_df = raw_alacak
                            # Dosya yolunu kendimiz tahmin ediyoruz
                            alacak_sonucu_path = os.path.join(paths["RAPOR"], f"{personel_adi}_ALACAK_RAPORU.xlsx")
                        
                        # Metrikleri Hesapla
                        toplam_fark = alacak_sonucu_df['Fark_FM_TL'].sum() + alacak_sonucu_df['Fark_UBGT_TL'].sum() + alacak_sonucu_df['Fark_HT_TL'].sum()
                        
                        st.session_state.metrikler = {
                            "alacak": f"{toplam_fark:,.2f} TL",
                            "mesai": f"{alacak_sonucu_df['Hak_FM_Saat'].sum():,.1f} Sa",
                            "ubgt": f"{alacak_sonucu_df['Hak_UBGT_Gun'].sum():,.1f} Gün",
                            "ht": f"{alacak_sonucu_df['Hak_HT_Gun'].sum():,.1f} Gün"
                        }
                        
                        st.session_state.hesaplama_tamam = True
                        status.update(label="✅ İşlem Başarıyla Tamamlandı!", state="complete", expanded=False)
                        time.sleep(1)
                        st.rerun() 
                        
                    except Exception as e:
                        st.error(f"Bir hata oluştu: {str(e)}")
                        # Detaylı hata analizi için:
                        import traceback
                        st.code(traceback.format_exc())

    # DURUM 3: Hesaplama bitmişse -> Sadece Sonuçları Göster
    elif st.session_state.hesaplama_tamam:
        
        st.toast("Hesaplama başarıyla tamamlandı. Raporlarınız aşağıdadır.", icon="✅")
        st.divider()
        
        col1, col2, col3, col4 = st.columns(4)
        m = st.session_state.metrikler
        col1.metric("Toplam Alacak", m["alacak"], delta="Fark Tutarı")
        col2.metric("Fazla Mesai", m["mesai"], delta="Toplam")
        col3.metric("UBGT", m["ubgt"], delta="Bayram")
        col4.metric("Hafta Tatili", m["ht"], delta="Pazar")

        st.subheader("📥 Raporları İndir")
        
        # Ekranı şu oranlarda bölüyoruz: [Boşluk(1) - Buton(2) - Buton(2) - Boşluk(1)]
        # Bu sayede butonlar ortada toplanır.
        bosluk_sol, col_btn1, col_btn2, bosluk_sag = st.columns([1, 2, 2, 1])
        
        # 1. ALACAK RAPORUNU BUL
        alacak_dosyalar = [f for f in os.listdir(paths["RAPOR"]) if f.endswith('.xlsx') and not f.startswith('~$')]
        if alacak_dosyalar:
            dosya_adi = alacak_dosyalar[0] # Klasörde zaten tek dosya olmalı
            alacak_yolu = os.path.join(paths["RAPOR"], dosya_adi)
            with open(alacak_yolu, "rb") as f:
                with col_btn1:
                    st.download_button(
                        f"💰 {"ALACAK RAPORU"}", # Buton üzerinde dosya adı görünsün
                        f, 
                        file_name=dosya_adi, 
                        key="dl_alacak_btn",
                        use_container_width=True
                    )

        # 2. PUANTAJ DOSYASINI BUL
        p_dosyalar = [f for f in os.listdir(paths["PUANTAJ"]) if f.endswith('.xlsx') and not f.startswith('~$')]
        if p_dosyalar:
            dosya_adi = p_dosyalar[0]
            p_yolu = os.path.join(paths["PUANTAJ"], dosya_adi)
            with open(p_yolu, "rb") as f:
                with col_btn2:
                    st.download_button(
                        f"📊 {"PUANTAJ"}", 
                        f, 
                        file_name=dosya_adi, 
                        key="dl_puan_btn",
                        use_container_width=True
                    )

        st.write("")
        col_sol, col_orta, col_sag = st.columns([1, 2, 1])
        with col_orta:
            if st.button("🔄 Yeni Hesaplama Yap", use_container_width=True):
                 st.session_state.hesaplama_tamam = False
                 st.session_state.ham_veri_onizleme = None
                 st.session_state.reset_counter += 1
                 st.rerun()

def auto_garbage_collector(max_age_seconds=1800): # 7200 saniye = 2 Saat
    """
    UserData klasöründeki eski oturum verilerini temizler.
    Kullanıcı sekmeyi kapatsa bile sunucuda yer açılmasını sağlar.
    """
    user_data_root = os.path.join(Config.BASE_DIR, "UserData")
    
    if not os.path.exists(user_data_root):
        return

    now = time.time()
    deleted_count = 0
    
    try:
        for session_dir in os.listdir(user_data_root):
            session_path = os.path.join(user_data_root, session_dir)
            
            # Klasörün son değiştirilme zamanına bakıyoruz
            st = os.stat(session_path)
            age = now - st.st_mtime
            
            # Eğer klasör belirlenen süreden daha eskiyse sil
            if age > max_age_seconds:
                shutil.rmtree(session_path)
                deleted_count += 1
        
        if deleted_count > 0:
            print(f"♻️ Garbage Collector: {deleted_count} eski oturum temizlendi.")
    except Exception as e:
        print(f"⚠️ Garbage Collector hatası: {e}")

# Uygulama her yüklendiğinde otomatik çalıştır
auto_garbage_collector()

if __name__ == "__main__":
    main()
